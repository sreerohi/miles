#!/usr/bin/env python3
"""
Analyze miles RL training log and output per-step metrics to Excel.

Usage:
    python analyze_training_log.py <log_file> [--steps N] [--warmup K] [--output FILE]

Produces an Excel file with three tabs:
  0) Rollout   - reward, truncated, lengths, log probs, gen throughput, token usage
  1) Train     - loss, KL, grad norm, throughput, TFLOPS, peak GPU memory
  2) Timers    - per-phase durations (seconds)
Each tab includes per-step rows plus average rows at the bottom.
"""

import re
import ast
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout


# ── ANSI escape code stripper ──────────────────────────────────────────────
ANSI_RE = re.compile(r'\x1b\[[0-9;]*m')

def strip_ansi(s: str) -> str:
    return ANSI_RE.sub('', s)


# ── Parsers for each metric category ──────────────────────────────────────

def parse_train_step_metrics(lines):
    """
    Parse train step dict lines from log_utils.py (deduplicate vs model.py).
    Returns dict: step_num -> {metric_name: value}
    """
    # Match: log_utils.py:414 - step N: {dict}
    pat = re.compile(r'log_utils\.py:\d+ - step (\d+): (\{.*\})')
    results = {}
    for line in lines:
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:  # take first occurrence only
                d = ast.literal_eval(m.group(2))
                results[step] = d
    return results


def parse_rollout_metrics(lines):
    """
    Parse rollout dict lines from log_utils.py.
    Returns dict: step_num -> {metric_name: value}
    """
    pat = re.compile(r'log_utils\.py:\d+ - rollout (\d+): (\{.*\})')
    results = {}
    for line in lines:
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:
                d = ast.literal_eval(m.group(2))
                results[step] = d
    return results


def parse_perf_train_metrics(lines):
    """
    Parse trainer perf dict from train_metric_utils.py log lines.
    Format: train_metric_utils.py:NN - perf N: {dict}
    Contains: perf/*_time, perf/*_tflops, perf/actor_train_tok_per_s,
              perf/step_time, perf/wait_time_ratio
    Returns dict: step_num -> {key: value}
    """
    pat = re.compile(r'train_metric_utils\.py:\d+ - perf (\d+): (\{.*\})')
    results = {}
    for line in lines:
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:
                results[step] = ast.literal_eval(m.group(2))
    return results


def parse_eval_metrics(lines):
    """
    Parse eval dict from rollout.py log lines (logged every --eval-interval steps).
    Format: rollout.py:NNN - eval N: {dict}
    Contains: eval/aime, eval/aime-pass@k, eval/aime/response_len/*, etc.
    Returns dict: eval_step -> {key: value}
    """
    pat = re.compile(r'rollout\.py:\d+ - eval (\d+): (\{.*\})')
    results = {}
    for line in lines:
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:
                results[step] = ast.literal_eval(m.group(2))
    return results


def parse_perf_rollout_metrics(lines):
    """
    Parse rollout perf dict from rollout.py log lines.
    Format: rollout.py:NN - perf N: {dict}
    Contains: perf/rollout_time, perf/tokens_per_gpu_per_sec,
              perf/longest_sample_tokens_per_sec, rollout/response_len/*, etc.
    Returns dict: step_num -> {key: value}
    """
    pat = re.compile(r'rollout\.py:\d+ - perf (\d+): (\{.*\})')
    results = {}
    for line in lines:
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:
                results[step] = ast.literal_eval(m.group(2))
    return results


def parse_gen_throughput(lines):
    """
    Parse gen throughput from SGLangEngine decode batch lines.
    Returns list of (line_index, throughput_value) tuples.
    """
    pat = re.compile(r'gen throughput \(token/s\): ([\d.]+)')
    results = []
    for i, line in enumerate(lines):
        m = pat.search(line)
        if m:
            results.append((i, float(m.group(1))))
    return results


def parse_rollout_line_indices(lines):
    """
    Find line indices of 'rollout N:' log lines (from log_utils.py).
    Returns dict: step_num -> line_index
    """
    pat = re.compile(r'log_utils\.py:\d+ - rollout (\d+):')
    results = {}
    for i, line in enumerate(lines):
        m = pat.search(line)
        if m:
            step = int(m.group(1))
            if step not in results:
                results[step] = i
    return results


def compute_gen_throughput_per_step(lines, max_steps):
    """
    Compute average gen throughput per step.
    Partition decode batch lines by rollout boundaries.
    Gen throughput lines BEFORE rollout 1 -> step 1
    Gen throughput lines BETWEEN rollout N and rollout N+1 -> step N+1
    """
    throughputs = parse_gen_throughput(lines)
    rollout_indices = parse_rollout_line_indices(lines)

    if not throughputs or not rollout_indices:
        return {}

    # Create sorted boundaries
    sorted_steps = sorted(rollout_indices.keys())
    boundaries = []  # list of (step, start_line_idx)
    for s in sorted_steps:
        boundaries.append((s, rollout_indices[s]))

    results = {}
    for step_idx, (step, boundary_line) in enumerate(boundaries):
        if step > max_steps:
            break

        # For step N: gen throughput lines are those BEFORE the rollout N line
        # and AFTER the previous rollout line (or start of file)
        if step_idx == 0:
            prev_boundary = 0
        else:
            prev_boundary = boundaries[step_idx - 1][1]

        step_throughputs = [
            tp for (li, tp) in throughputs
            if li >= prev_boundary and li < boundary_line
        ]

        if step_throughputs:
            # Skip the first measurement (often artificially low due to warmup)
            if len(step_throughputs) > 1 and step_throughputs[0] < 1000:
                step_throughputs = step_throughputs[1:]
            results[step] = {
                'avg': sum(step_throughputs) / len(step_throughputs),
                'max': max(step_throughputs),
                'min': min(step_throughputs),
            }

    return results


def parse_memory_per_step(lines):
    """
    Parse memory at 'before offload model' (peak training memory).
    Uses any rank's data (Rank 0 is often suppressed by Ray's log deduplication).
    Groups by step using 'Timer train end' boundaries.
    Returns dict: step_num -> {used_GB, free_GB, allocated_GB, reserved_GB}
    """
    mem_pat = re.compile(
        r'Memory-Usage before offload model: '
        r"\{'gpu': '\d+', 'total_GB': ([\d.]+), 'free_GB': ([\d.]+), "
        r"'used_GB': ([\d.]+), 'allocated_GB': ([\d.]+), 'reserved_GB': ([\d.]+)\}"
    )
    train_end_pat = re.compile(r'Timer train end \(elapsed:')

    results = {}
    step = 1
    current_mem = None
    for line in lines:
        m = mem_pat.search(line)
        if m:
            # Take the first "before offload" memory in each step (peak memory)
            if current_mem is None:
                current_mem = {
                    'total_GB': float(m.group(1)),
                    'free_GB': float(m.group(2)),
                    'used_GB': float(m.group(3)),
                    'allocated_GB': float(m.group(4)),
                    'reserved_GB': float(m.group(5)),
                }
        if train_end_pat.search(line):
            if current_mem is not None:
                results[step] = current_mem
                current_mem = None
            step += 1

    return results


def parse_step_wall_times(lines):
    """
    Compute per-step wall-clock time from timestamps on 'Timer train end' lines.
    Step N wall time = timestamp(train end N) - timestamp(train end N-1).
    Step 1 has no predecessor, so it gets no wall time.
    Returns dict: step_num -> wall_time_seconds
    """
    pat = re.compile(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\].*Timer train end')
    timestamps = []
    for line in lines:
        m = pat.search(line)
        if m:
            ts = datetime.strptime(m.group(1), '%Y-%m-%d %H:%M:%S')
            timestamps.append(ts)

    results = {}
    for i in range(1, len(timestamps)):
        step = i + 1  # step 2 = delta between train end 1 and train end 2
        delta = (timestamps[i] - timestamps[i - 1]).total_seconds()
        results[step] = delta

    return results


def parse_token_usage_per_step(lines):
    """
    Parse max token usage from decode batch lines per step.
    Token usage field: 'token usage: 0.05'
    Partitioned by rollout boundaries.
    """
    pat = re.compile(r'token usage: ([\d.]+)')
    rollout_indices = parse_rollout_line_indices(lines)

    if not rollout_indices:
        return {}

    sorted_steps = sorted(rollout_indices.keys())
    boundaries = [(s, rollout_indices[s]) for s in sorted_steps]

    # Collect all token usage values with their line indices
    usages = []
    for i, line in enumerate(lines):
        if 'SGLangEngine' in line and 'token usage:' in line:
            m = pat.search(line)
            if m:
                usages.append((i, float(m.group(1))))

    results = {}
    for step_idx, (step, boundary_line) in enumerate(boundaries):
        if step_idx == 0:
            prev_boundary = 0
        else:
            prev_boundary = boundaries[step_idx - 1][1]

        step_usages = [u for (li, u) in usages if li >= prev_boundary and li < boundary_line]
        if step_usages:
            results[step] = max(step_usages)

    return results


# ── Excel output ──────────────────────────────────────────────────────────

# Number format helpers: choose sensible decimal places per metric type
def _excel_num_fmt(fmt_hint):
    """Convert our format hint to an Excel number format string."""
    if fmt_hint == "int":
        return '0'
    if fmt_hint == "sci":
        return '0.00E+00'
    # e.g. "2" -> "0.00", "3" -> "0.000"
    try:
        d = int(fmt_hint)
        return '0.' + '0' * d
    except ValueError:
        return '0.00'


def _write_sheet(wb, sheet_name, columns, all_step_data, max_steps, warmup_steps, is_first=False, stride=1):
    """
    Write one tab with per-step rows + average rows at the bottom.

    columns: list of (display_header, data_key, fmt_hint)
        fmt_hint: "int", "sci", or a digit string like "2" for 2 decimal places.
    stride: only display every stride-th step (averages are still over all steps).
    """
    if is_first:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    avg_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    avg_fill_warm = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    avg_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )
    thick_top = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='medium'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    steps = [s for s in sorted(all_step_data.keys()) if s <= max_steps]
    summary_steps = [s for s in steps if s > warmup_steps]
    display_steps = steps[::stride]  # thinned for display; avg rows still use all steps

    # ── Header row ────────────────────────────────────────────────────────
    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, step in enumerate(display_steps, 2):
        data = all_step_data[step]
        # Alternate row shading for readability
        row_fill = None
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col_idx, (header, key, fmt) in enumerate(columns, 1):
            val = data.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is not None:
                cell.value = val
                cell.number_format = _excel_num_fmt(fmt)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if row_fill:
                cell.fill = row_fill

    # ── Average rows ──────────────────────────────────────────────────────
    avg_row_all = len(display_steps) + 2
    avg_row_excl = len(display_steps) + 3

    for avg_row, label, step_subset, fill in [
        (avg_row_all, f"Mean (all {max_steps} steps)", steps, avg_fill),
        (avg_row_excl, f"Mean (steps {warmup_steps+1}-{max_steps})", summary_steps, avg_fill_warm),
    ]:
        for col_idx, (header, key, fmt) in enumerate(columns, 1):
            cell = ws.cell(row=avg_row, column=col_idx)
            cell.border = thick_top if col_idx == 1 else thin_border
            cell.border = thick_top
            cell.fill = fill
            cell.font = avg_font

            if key == "step":
                cell.value = label
                cell.alignment = Alignment(horizontal='left')
                continue

            vals = [all_step_data[s][key] for s in step_subset
                    if key in all_step_data[s] and all_step_data[s][key] is not None]
            if vals:
                cell.value = sum(vals) / len(vals)
                cell.number_format = _excel_num_fmt(fmt)
            cell.alignment = Alignment(horizontal='right')

    # ── Column widths ─────────────────────────────────────────────────────
    for col_idx, (header, key, fmt) in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        # Width based on header length, with a reasonable min/max
        w = max(len(header) + 3, 10)
        if key == "step":
            w = max(w, 30)  # room for avg labels
        ws.column_dimensions[col_letter].width = min(w, 35)

    ws.freeze_panes = 'B2'
    return ws


def write_excel(all_step_data, all_eval_data, max_steps, warmup_steps, output_path, stride=1):
    wb = openpyxl.Workbook()

    # ── Tab 0: Eval ────────────────────────────────────────────────────────
    if all_eval_data:
        eval_cols = [
            ("Step",             "step",                              "int"),
            ("AIME Accuracy",    "eval/aime",                         "3"),
            ("Pass@1",           "eval/aime-pass@1",                  "3"),
            ("Pass@4",           "eval/aime-pass@4",                  "3"),
            ("Pass@8",           "eval/aime-pass@8",                  "3"),
            ("Pass@16",          "eval/aime-pass@16",                 "3"),
            ("Truncated",        "eval/aime-truncated_ratio",         "3"),
            ("Resp Len Mean",    "eval/aime/response_len/mean",       "0"),
        ]
        max_eval_step = max(all_eval_data.keys())
        _write_sheet(wb, "Eval", eval_cols, all_eval_data, max_eval_step, 0, is_first=True)

    # ── Tab 1: Rollout ─────────────────────────────────────────────────────
    # Includes rollout generation throughput & token usage (from SGLangEngine)
    rollout_cols = [
        ("Step",                    "step",                          "int"),
        ("Rollout Tok/s",           "gen_throughput_avg",            "0"),
        ("Rollout Tok/GPU/s",       "perf/tokens_per_gpu_per_sec",   "0"),
        ("Raw Reward",              "rollout/raw_reward",            "3"),
        ("Truncated",               "rollout/truncated",             "3"),
        ("Max Token Usage",         "max_token_usage",               "3"),
    ]
    is_rollout_first = not all_eval_data
    _write_sheet(wb, "Rollout", rollout_cols, all_step_data, max_steps, warmup_steps, is_first=is_rollout_first, stride=stride)

    # ── Tab 1: Train ────────────────────────────────────────────────────
    # Includes peak GPU memory (captured right after training, before offload)
    train_cols = [
        ("Step",                "step",                          "int"),
        ("Train Tok/s",         "perf/actor_train_tok_per_s",    "0"),
        ("Train TFLOPS",        "perf/actor_train_tflops",       "1"),
        ("Peak Used (GB)",      "mem/used_GB",                   "2"),
        ("Peak Free (GB)",      "mem/free_GB",                   "2"),
        ("Grad Norm",           "train/grad_norm",               "4"),
        ("PPO KL",              "train/ppo_kl",                  "sci"),
        ("PG Clipfrac",         "train/pg_clipfrac",             "3"),
    ]
    _write_sheet(wb, "Train", train_cols, all_step_data, max_steps, warmup_steps, stride=stride)

    # ── Tab 2: Timers ───────────────────────────────────────────────────
    timer_cols = [
        ("Step",                "step",                         "int"),
        ("Step Time (s)",       "perf/step_time",              "1"),
        ("Rollout (s)",         "perf/rollout_time",            "1"),
        ("Rollout in Wait (s)", "timer/rollout_in_wait_time",   "1"),
        ("Train Total (s)",     "perf/train_time",              "1"),
        ("Train Other (s)",     "timer/train_other_time",       "1"),
        ("Trainer Wait (s)",    "perf/train_wait_time",         "1"),
        ("Actor Train (s)",     "perf/actor_train_time",        "1"),
        ("Ref Log Probs (s)",   "perf/ref_log_probs_time",      "1"),
        ("Log Probs (s)",       "perf/log_probs_time",          "1"),
        ("Sleep (s)",           "perf/sleep_time",              "1"),
        ("Update Weights (s)",  "perf/update_weights_time",     "1"),
        ("Save Model (s)",      "perf/save_model_time",         "1"),
        ("Wake Up (s)",         "perf/wake_up_time",            "1"),
        ("Data Preprocess (s)", "perf/data_preprocess_time",    "1"),
        ("Wait Other (s)",      "timer/wait_other_time",        "1"),
    ]
    ws_timers = _write_sheet(wb, "Timers", timer_cols, all_step_data, max_steps, warmup_steps, stride=stride)

    # ── Pie chart: true partition of step time (excl. warmup mean) ──────
    # Non-overlapping decomposition:
    #   step_time = train_time + train_wait_time
    #   train_time = actor_train + ref_log_probs + log_probs + train_other
    #   train_wait_time = rollout + sleep + update_weights + save_model + wake_up + data_preprocess + wait_other
    # This guarantees no overlap and no missing time.
    summary_steps_excel = [s for s in sorted(all_step_data.keys()) if s <= max_steps and s > warmup_steps]
    partition_steps = [
        s for s in summary_steps_excel
        if all_step_data[s].get("perf/train_time") is not None and all_step_data[s].get("perf/train_wait_time") is not None
    ]

    def _safe_nonneg(data, key):
        val = data.get(key)
        if isinstance(val, (int, float)):
            return max(float(val), 0.0)
        return 0.0

    partition_labels = [
        "Actor Train",
        "Ref Log Probs",
        "Log Probs",
        "Train Other",
        "Rollout",
        "Sleep",
        "Update Weights",
        "Save Model",
        "Wake Up",
        "Data Preprocess",
        "Wait Other",
    ]
    partition_sums = {label: 0.0 for label in partition_labels}

    for s in partition_steps:
        d = all_step_data[s]
        train_total = _safe_nonneg(d, "perf/train_time")
        wait_total = _safe_nonneg(d, "perf/train_wait_time")

        actor_train = _safe_nonneg(d, "perf/actor_train_time")
        ref_log_probs = _safe_nonneg(d, "perf/ref_log_probs_time")
        log_probs = _safe_nonneg(d, "perf/log_probs_time")
        train_known = actor_train + ref_log_probs + log_probs
        if train_known > train_total and train_known > 0:
            # Guard against tiny logging mismatch; keep strict partition.
            scale = train_total / train_known
            actor_train *= scale
            ref_log_probs *= scale
            log_probs *= scale
            train_known = train_total
        train_other = train_total - train_known

        sleep_time = _safe_nonneg(d, "perf/sleep_time")
        update_weights = _safe_nonneg(d, "perf/update_weights_time")
        save_model_time = _safe_nonneg(d, "perf/save_model_time")
        wake_up = _safe_nonneg(d, "perf/wake_up_time")
        data_preprocess = _safe_nonneg(d, "perf/data_preprocess_time")
        wait_known = sleep_time + update_weights + save_model_time + wake_up + data_preprocess
        if wait_known > wait_total and wait_known > 0:
            # Guard against tiny logging mismatch; keep strict partition.
            scale = wait_total / wait_known
            sleep_time *= scale
            update_weights *= scale
            save_model_time *= scale
            wake_up *= scale
            data_preprocess *= scale
            wait_known = wait_total

        # rollout_time is measured by rollout manager; clamp it to remaining wait budget
        # so the pie remains an exact partition with no overlap.
        rollout_time = min(_safe_nonneg(d, "perf/rollout_time"), max(wait_total - wait_known, 0.0))
        wait_other = wait_total - wait_known - rollout_time

        partition_sums["Actor Train"] += actor_train
        partition_sums["Ref Log Probs"] += ref_log_probs
        partition_sums["Log Probs"] += log_probs
        partition_sums["Train Other"] += train_other
        partition_sums["Rollout"] += rollout_time
        partition_sums["Sleep"] += sleep_time
        partition_sums["Update Weights"] += update_weights
        partition_sums["Save Model"] += save_model_time
        partition_sums["Wake Up"] += wake_up
        partition_sums["Data Preprocess"] += data_preprocess
        partition_sums["Wait Other"] += wait_other

    partition_means = {
        label: (partition_sums[label] / len(partition_steps) if partition_steps else 0.0)
        for label in partition_labels
    }
    # Sort largest slice first so the pie chart reads naturally.
    partition_labels = sorted(partition_labels, key=lambda l: partition_means[l], reverse=True)
    total_mean = sum(partition_means.values())

    # Write a small helper table to the right of the main data (col 18-19)
    # so the chart can reference live cell values.
    HC = 18  # helper start column (R)
    HR = 2   # helper start row (row 1 is the frozen header)
    header_style = Font(bold=True)
    ws_timers.cell(row=HR, column=HC, value="Phase").font = header_style
    ws_timers.cell(
        row=HR,
        column=HC + 1,
        value=f"Mean (s) steps {warmup_steps+1}-{max_steps}",
    ).font = header_style

    for i, label in enumerate(partition_labels):
        row = HR + 1 + i
        ws_timers.cell(row=row, column=HC, value=label)
        val_cell = ws_timers.cell(row=row, column=HC + 1, value=partition_means[label])
        val_cell.number_format = "0.00"

    total_row = HR + 1 + len(partition_labels)
    ws_timers.cell(row=total_row, column=HC, value="Total (Step Time)").font = Font(bold=True)
    total_cell = ws_timers.cell(row=total_row, column=HC + 1, value=total_mean)
    total_cell.font = Font(bold=True)
    total_cell.number_format = "0.00"

    ws_timers.column_dimensions[get_column_letter(HC)].width = 18
    ws_timers.column_dimensions[get_column_letter(HC + 1)].width = 28

    # Build and place the pie chart
    n = len(partition_labels)
    labels_ref = Reference(ws_timers, min_col=HC, min_row=HR + 1, max_row=HR + n)
    data_ref = Reference(ws_timers, min_col=HC + 1, min_row=HR, max_row=HR + n)
    chart = PieChart()
    chart.title = "Timer Step-Time Partition (No Overlap)"
    chart.style = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.legend.position = "b"  # legend below pie
    chart.width = 15  # cm
    chart.height = 18  # cm
    ws_timers.add_chart(chart, get_column_letter(HC + 3) + "2")

    wb.save(output_path)
    print(f"Saved to {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Analyze miles RL training log")
    parser.add_argument("log_file", help="Path to the log file")
    parser.add_argument("--steps", type=int, default=15, help="Number of steps to analyze (default: 15)")
    parser.add_argument("--warmup", type=int, default=3, help="Number of warmup steps to exclude from summary mean (default: 3)")
    parser.add_argument("--stride", type=int, default=1, help="Display every k-th step in Rollout/Train/Timers tables (default: 1); averages always over all steps")
    parser.add_argument("--output", type=str, default=None, help="Output Excel file path")
    args = parser.parse_args()

    log_path = Path(args.log_file)
    if args.output:
        output_path = args.output
    else:
        output_path = log_path.with_suffix('.xlsx')

    max_steps = args.steps
    warmup_steps = args.warmup
    stride = args.stride

    print(f"Reading {log_path}...")
    with open(log_path, 'r', errors='replace') as f:
        raw_lines = f.readlines()

    # Strip ANSI codes
    lines = [strip_ansi(line) for line in raw_lines]
    print(f"  {len(lines)} lines read, ANSI codes stripped.")

    # Parse all metric categories
    print("Parsing train step metrics...")
    train_metrics = parse_train_step_metrics(lines)
    print(f"  Found {len(train_metrics)} steps")

    print("Parsing rollout metrics...")
    rollout_metrics = parse_rollout_metrics(lines)
    print(f"  Found {len(rollout_metrics)} rollouts")

    print("Parsing trainer perf metrics...")
    perf_train_metrics = parse_perf_train_metrics(lines)
    print(f"  Found trainer perf for {len(perf_train_metrics)} steps")

    print("Parsing rollout perf metrics...")
    perf_rollout_metrics = parse_perf_rollout_metrics(lines)
    print(f"  Found rollout perf for {len(perf_rollout_metrics)} steps")

    print("Parsing gen throughput...")
    gen_throughput = compute_gen_throughput_per_step(lines, max_steps)
    print(f"  Found gen throughput for {len(gen_throughput)} steps")

    print("Parsing memory usage...")
    memory_metrics = parse_memory_per_step(lines)
    print(f"  Found memory data for {len(memory_metrics)} steps")

    print("Parsing token usage...")
    token_usage = parse_token_usage_per_step(lines)
    print(f"  Found token usage for {len(token_usage)} steps")

    print("Parsing step wall times...")
    step_wall_times = parse_step_wall_times(lines)
    print(f"  Found wall times for {len(step_wall_times)} steps")

    print("Parsing eval metrics...")
    eval_metrics_raw = parse_eval_metrics(lines)
    print(f"  Found eval data for {len(eval_metrics_raw)} eval steps")

    # Build eval data (sparse — one entry per eval interval)
    all_eval_data = {}
    for step, d in eval_metrics_raw.items():
        entry = {"step": step}
        entry.update(d)
        all_eval_data[step] = entry

    # Merge all metrics per step
    all_step_data = {}
    for step in range(1, max_steps + 1):
        data = {"step": step}

        # Train metrics
        if step in train_metrics:
            for k, v in train_metrics[step].items():
                data[k] = v

        # Rollout metrics
        if step in rollout_metrics:
            for k, v in rollout_metrics[step].items():
                data[k] = v

        # Trainer perf dict (timings, tflops, tok/s)
        if step in perf_train_metrics:
            for k, v in perf_train_metrics[step].items():
                data[k] = v

        # Rollout perf dict (rollout_time, tokens_per_gpu_per_sec, etc.)
        if step in perf_rollout_metrics:
            for k, v in perf_rollout_metrics[step].items():
                data[k] = v

        # Gen throughput from SGLangEngine decode batch lines
        if step in gen_throughput:
            data["gen_throughput_avg"] = gen_throughput[step]['avg']
            data["gen_throughput_max"] = gen_throughput[step]['max']
            data["gen_throughput_min"] = gen_throughput[step]['min']

        # Token usage
        if step in token_usage:
            data["max_token_usage"] = token_usage[step]

        # Memory
        if step in memory_metrics:
            for k, v in memory_metrics[step].items():
                data[f"mem/{k}"] = v

        # Step wall time (from timestamp deltas)
        if step in step_wall_times:
            data["step_wall_time"] = step_wall_times[step]

        # Always expose save_model_time as 0.0 when not present for this step.
        data["perf/save_model_time"] = max(float(data.get("perf/save_model_time") or 0.0), 0.0)

        # Derived: train_wait / train ratio
        tw = data.get("perf/train_wait_time")
        tt = data.get("perf/train_time")
        if tw is not None and tt is not None and tt > 0:
            data["train_wait_ratio"] = tw / tt

        # Derived: non-overlapping timer partition terms.
        # This ensures no missing time in step-level accounting:
        #   step_time = train_time + train_wait_time
        #   train_time = actor_train + ref_log_probs + log_probs + train_other
        #   train_wait_time = sleep + update_weights + save_model + wake_up + data_preprocess + rollout_in_wait + wait_other
        if tw is not None and tt is not None:
            tw_v = max(float(tw), 0.0)
            tt_v = max(float(tt), 0.0)
            data["perf/step_time"] = tt_v + tw_v

            actor_train = max(float(data.get("perf/actor_train_time") or 0.0), 0.0)
            ref_log_probs = max(float(data.get("perf/ref_log_probs_time") or 0.0), 0.0)
            log_probs = max(float(data.get("perf/log_probs_time") or 0.0), 0.0)
            train_known = actor_train + ref_log_probs + log_probs
            if train_known > tt_v and train_known > 0:
                scale = tt_v / train_known
                actor_train *= scale
                ref_log_probs *= scale
                log_probs *= scale
                train_known = tt_v
            data["timer/train_other_time"] = tt_v - train_known

            sleep_time = max(float(data.get("perf/sleep_time") or 0.0), 0.0)
            update_weights = max(float(data.get("perf/update_weights_time") or 0.0), 0.0)
            save_model_time = max(float(data.get("perf/save_model_time") or 0.0), 0.0)
            wake_up = max(float(data.get("perf/wake_up_time") or 0.0), 0.0)
            data_preprocess = max(float(data.get("perf/data_preprocess_time") or 0.0), 0.0)
            wait_known = sleep_time + update_weights + save_model_time + wake_up + data_preprocess
            if wait_known > tw_v and wait_known > 0:
                scale = tw_v / wait_known
                sleep_time *= scale
                update_weights *= scale
                save_model_time *= scale
                wake_up *= scale
                data_preprocess *= scale
                wait_known = tw_v

            rollout_raw = max(float(data.get("perf/rollout_time") or 0.0), 0.0)
            rollout_in_wait = min(rollout_raw, max(tw_v - wait_known, 0.0))
            data["timer/rollout_in_wait_time"] = rollout_in_wait
            data["timer/wait_other_time"] = tw_v - wait_known - rollout_in_wait

        all_step_data[step] = data

    # Print summary to console
    def fmt_val(v, decimals=4):
        if v is None or v == 'N/A':
            return 'N/A'
        if isinstance(v, float):
            if abs(v) < 1e-5 and v != 0:
                return f"{v:.2e}"
            return f"{v:.{decimals}f}"
        return str(v)

    def print_avg_row(label, step_subset, keys_fmts, widths):
        """Print a formatted average row for the given steps and (key, decimals) pairs."""
        parts = [f"{label:>{widths[0]}}"]
        for i, (key, dec) in enumerate(keys_fmts):
            vals = [all_step_data[s][key] for s in step_subset
                    if key in all_step_data[s] and all_step_data[s][key] is not None]
            if vals:
                avg = sum(vals) / len(vals)
                parts.append(f"{fmt_val(avg, dec):>{widths[i+1]}}")
            else:
                parts.append(f"{'N/A':>{widths[i+1]}}")
        print("  ".join(parts))

    steps = list(range(1, max_steps + 1))
    summary_steps = [s for s in steps if s > warmup_steps]
    display_steps = steps[::stride]

    print(f"\n{'='*130}")
    print(f"Per-Step Metrics (steps 1-{max_steps}, warmup={warmup_steps})")
    print(f"{'='*130}")

    # ── Eval ─────────────────────────────────────────────────────────────────
    if all_eval_data:
        eval_steps = sorted(all_eval_data.keys())
        print(f"\n--- Eval (AIME, every {eval_steps[1]-eval_steps[0] if len(eval_steps)>1 else '?'} steps) ---")
        e_hdr = f"{'Step':>4}  {'aime_acc':>8}  {'pass@1':>6}  {'pass@4':>6}  {'pass@8':>6}  {'pass@16':>7}  {'trunc':>6}  {'resp_len':>8}"
        print(e_hdr)
        print("-" * len(e_hdr))
        for step in eval_steps:
            d = all_eval_data[step]
            print(f"{step:4d}  "
                  f"{fmt_val(d.get('eval/aime'), 3):>8}  "
                  f"{fmt_val(d.get('eval/aime-pass@1'), 3):>6}  "
                  f"{fmt_val(d.get('eval/aime-pass@4'), 3):>6}  "
                  f"{fmt_val(d.get('eval/aime-pass@8'), 3):>6}  "
                  f"{fmt_val(d.get('eval/aime-pass@16'), 3):>7}  "
                  f"{fmt_val(d.get('eval/aime-truncated_ratio'), 3):>6}  "
                  f"{fmt_val(d.get('eval/aime/response_len/mean'), 0):>8}")

    # ── Rollout ──────────────────────────────────────────────────────────────
    print(f"\n--- Rollout ---")
    r_hdr = f"{'Step':>4}  {'rl_tok/s':>8}  {'rl_tok/gpu/s':>12}  {'raw_rwd':>8}  {'trunc':>6}  {'tok_usage':>9}"
    print(r_hdr)
    print("-" * len(r_hdr))
    for step in display_steps:
        d = all_step_data[step]
        print(f"{step:4d}  "
              f"{fmt_val(d.get('gen_throughput_avg'), 0):>8}  "
              f"{fmt_val(d.get('perf/tokens_per_gpu_per_sec'), 0):>12}  "
              f"{fmt_val(d.get('rollout/raw_reward'), 3):>8}  "
              f"{fmt_val(d.get('rollout/truncated'), 3):>6}  "
              f"{fmt_val(d.get('max_token_usage'), 3):>9}")
    r_keys = [("gen_throughput_avg", 0), ("perf/tokens_per_gpu_per_sec", 0),
              ("rollout/raw_reward", 3), ("rollout/truncated", 3), ("max_token_usage", 3)]
    r_w = [4, 8, 12, 8, 6, 9]
    print("-" * len(r_hdr))
    print_avg_row(f"avg(1-{max_steps})", steps, r_keys, r_w)
    print_avg_row(f"avg({warmup_steps+1}-{max_steps})", summary_steps, r_keys, r_w)

    # ── Train ────────────────────────────────────────────────────────────────
    print(f"\n--- Train ---")
    t_hdr = f"{'Step':>4}  {'trn_tok/s':>9}  {'trn_tflops':>10}  {'used_GB':>8}  {'free_GB':>8}  {'grad_norm':>9}  {'ppo_kl':>10}  {'clipfrac':>8}"
    print(t_hdr)
    print("-" * len(t_hdr))
    for step in display_steps:
        d = all_step_data[step]
        print(f"{step:4d}  "
              f"{fmt_val(d.get('perf/actor_train_tok_per_s'), 0):>9}  "
              f"{fmt_val(d.get('perf/actor_train_tflops'), 1):>10}  "
              f"{fmt_val(d.get('mem/used_GB'), 2):>8}  "
              f"{fmt_val(d.get('mem/free_GB'), 2):>8}  "
              f"{fmt_val(d.get('train/grad_norm'), 4):>9}  "
              f"{fmt_val(d.get('train/ppo_kl')):>10}  "
              f"{fmt_val(d.get('train/pg_clipfrac'), 3):>8}")
    t_keys = [("perf/actor_train_tok_per_s", 0), ("perf/actor_train_tflops", 1),
              ("mem/used_GB", 2), ("mem/free_GB", 2),
              ("train/grad_norm", 4), ("train/ppo_kl", 4), ("train/pg_clipfrac", 3)]
    t_w = [4, 9, 10, 8, 8, 9, 10, 8]
    print("-" * len(t_hdr))
    print_avg_row(f"avg(1-{max_steps})", steps, t_keys, t_w)
    print_avg_row(f"avg({warmup_steps+1}-{max_steps})", summary_steps, t_keys, t_w)

    # ── Timers (per-phase durations) ──────────────────────────────────────
    print(f"\n--- Timers (seconds) ---")
    tm_hdr = (
        f"{'Step':>4}  {'step':>7}  {'rollout':>8}  {'roll_in_w':>9}  "
        f"{'train':>7}  {'trn_oth':>8}  {'trainer_wait':>12}  {'actor_trn':>9}  "
        f"{'ref_logp':>8}  {'log_probs':>9}  {'sleep':>6}  {'upd_wts':>7}  {'save':>6}  "
        f"{'wake_up':>7}  {'data_pp':>7}  {'wait_oth':>8}"
    )
    print(tm_hdr)
    print("-" * len(tm_hdr))
    for step in display_steps:
        d = all_step_data[step]
        print(f"{step:4d}  {fmt_val(d.get('perf/step_time'), 1):>7}  "
              f"{fmt_val(d.get('perf/rollout_time'), 1):>8}  "
              f"{fmt_val(d.get('timer/rollout_in_wait_time'), 1):>9}  "
              f"{fmt_val(d.get('perf/train_time'), 1):>7}  "
              f"{fmt_val(d.get('timer/train_other_time'), 1):>8}  "
              f"{fmt_val(d.get('perf/train_wait_time'), 1):>12}  "
              f"{fmt_val(d.get('perf/actor_train_time'), 1):>9}  "
              f"{fmt_val(d.get('perf/ref_log_probs_time'), 1):>8}  "
              f"{fmt_val(d.get('perf/log_probs_time'), 1):>9}  "
              f"{fmt_val(d.get('perf/sleep_time'), 1):>6}  "
              f"{fmt_val(d.get('perf/update_weights_time'), 1):>7}  "
              f"{fmt_val(d.get('perf/save_model_time'), 1):>6}  "
              f"{fmt_val(d.get('perf/wake_up_time'), 1):>7}  "
              f"{fmt_val(d.get('perf/data_preprocess_time'), 1):>7}  "
              f"{fmt_val(d.get('timer/wait_other_time'), 1):>8}")
    tm_keys = [("perf/step_time", 1), ("perf/rollout_time", 1), ("timer/rollout_in_wait_time", 1),
               ("perf/train_time", 1), ("timer/train_other_time", 1), ("perf/train_wait_time", 1),
               ("perf/actor_train_time", 1), ("perf/ref_log_probs_time", 1), ("perf/log_probs_time", 1),
               ("perf/sleep_time", 1), ("perf/update_weights_time", 1), ("perf/save_model_time", 1), ("perf/wake_up_time", 1),
               ("perf/data_preprocess_time", 1), ("timer/wait_other_time", 1)]
    tm_w = [4, 7, 8, 9, 7, 8, 12, 9, 8, 9, 6, 7, 6, 7, 7, 8]
    print("-" * len(tm_hdr))
    print_avg_row(f"avg(1-{max_steps})", steps, tm_keys, tm_w)
    print_avg_row(f"avg({warmup_steps+1}-{max_steps})", summary_steps, tm_keys, tm_w)

    # Write Excel
    print(f"\nWriting Excel to {output_path}...")
    write_excel(all_step_data, all_eval_data, max_steps, warmup_steps, output_path, stride=stride)
    print("Done!")


if __name__ == "__main__":
    main()
